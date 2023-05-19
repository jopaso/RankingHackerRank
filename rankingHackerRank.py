from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side

class Problem_Solved:
    def __init__(self,name, time):
        self.name = name
        self.time = time


class Team:
    def __init__(self, name):
        self.name = name
        self.problems_solved = []
        self.problems_failed = []
        self.time = 0

    def solve(self, problem, time):
        if problem not in self.problems_solved:
            self.problems_solved.append(Problem_Solved(problem, time))
            self.time += int(time)

            if problem in self.problems_failed:
                self.problems_failed.remove(problem)

    def fail(self, problem):
        if problem in self.problems_solved or problem in self.problems_failed:
            return 0
        else:
            self.problems_failed.append(problem)
    
    def punctuation(self):
        return len(self.problems_solved)
    
    def get(self, name):
        for i in self.problems_solved:
            if i.name == name:
                return i

    def __eq__(self, other):
        return self.punctuation() == other.punctuation() and self.time == other.time
    
    def __gt__(self, other):
        if self.punctuation() == other.punctuation():
            return self.time < other.time
        return self.punctuation() > other.punctuation()

    def __lt__(self, other):
        if self.punctuation() == other.punctuation():
            return self.time > other.time
        return self.punctuation() <  other.punctuation()





def mostrar(submissions):
    resultados = submissions.copy()
    
    #Delete useless information
    for res in resultados:
        res.pop(2) #Deleting ID
        res.pop(2) #Deleting Language
        res.pop(4) #Deleting Score
        res.pop(5) #Deleting View

    with open('./submissions_Accepted.txt', 'w') as f:
        #Show the results in a table
        res = "{:<30}".format('PROBLEM') + "{:<30}".format('TEAM') + "{:<5}".format('TIME')

        res += '\n' + ('-' * 70)
        for i in resultados:
            res += "\n{:<30}".format(i[0]) + "{:<30}".format(i[1]) + "{:<5}".format(i[2])
            res += '\n' + ('-' * 70)    
        
        print(res)
        f.write(res + '|n')

def ini_excel():

    # Create a new excel book
    book = Workbook()

    # Seleccionar la hoja de trabajo activa
    hoja = book.active

    # Agregar datos a la hoja de trabajo
    hoja['A1'] = 'Equipo/Problemas'
    hoja['B1'] = 'Hackear cuentas'
    hoja['C1'] = 'Contando bancos'
    hoja['D1'] = 'El banco perfecto'
    hoja['E1'] = 'La caja fuerte'
    hoja['F1'] = 'La banda'
    hoja['G1'] = 'Guardias de seguridad'
    hoja['H1'] = 'Cajeros'
    hoja['I1'] = 'Cámaras'
    hoja['J1'] = '¡A por el botín!'

    letras = 'ABCDEFGHIJ'
    for l in letras:
        hoja.column_dimensions[l].width = 20

    hoja.row_dimensions[1].height = 10
    return book

def escribe_excel(team, position, book):
    letras = 'BCDEFGHIJ'
    hoja = book.active
    hoja['A' + str(position + 2)] = team.name

    for l in letras:
        prob = team.get(hoja[l + '1'].value ) #go accross all the problems and check if it has been solved
        cell = hoja[l + str(position + 2)]
        hoja.row_dimensions[position + 2].height = 30

        if prob is None:
            #cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid') #background green
            print(f'{hoja[l + "1"].value} NOT Solved')
        else:
            cell.value = prob.time
            #cell.value = prob.time
            cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid') #background green
            cell.border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
            print(f'{hoja[l + "1"].value} Solved')
        


def main():
    with open('Input.in','r') as inp:
        data = inp.read().strip().split('\n\n')
        #print(data)

    #Put the submissions in a matrix where each row represents a submission
    submissions =[]
    j = 0
    for i in range (len(data) // 9):
        submissions.append(data[j:j+9])
        
        j += 9

    #Remove out of time submissions
    i = 0
    while i < len(submissions):
        if 'No' in submissions[i][7]:
            submissions.pop(i)
        else:
            i += 1
    
    for i in submissions:
        print(i)

    print()
    
    mostrar(submissions)

    teams = {} #key:value -> Team name : team object
    
    for sub in submissions:
        if teams.get(sub[1]) is None: #If the teams hasn't already appeared, add it to the dict
            teams[sub[1]] = Team(sub[1])
        
        if 'Accepted' in sub[3]:
            teams.get(sub[1]).solve(sub[0], sub[2])
        else:
            teams.get(sub[1]).fail(sub[0])
   
    with open('./clasificacion.txt', 'w') as f:
        print('CLASIFICACION:')
        # Print the ranking
        teams_sorted = sorted(teams.values(), reverse=True)
        
        book = ini_excel()

        for posicion, t in enumerate(teams_sorted):
            print(f'{posicion + 1}. {t.name} -->  {t.punctuation()} problemas resueltos // tiempo = {t.time}')
            f.write(f'{posicion + 1}. {t.name} -->  {t.punctuation()} problemas resueltos // tiempo = {t.time}\n')
            escribe_excel(t, posicion, book)
            posicion += 1

        book.save('resultados.xlsx')

if __name__ == '__main__':
    main()



